'''
Created on 1 Oct. 2018

@author: Mykhailo Savytskyi
@email: m.savytskyi@unsw.edu.au
'''
import csv
from scipy.interpolate import interp1d
import matplotlib.pyplot as plt
from os import listdir
from os.path import isfile, join
from fnmatch import fnmatch


def getFiles(folder):

    onlyfiles = [f for f in listdir(folder) if isfile(join(folder, f))]
    return onlyfiles


def findFiles(folder, pattern):
    files = getFiles(folder)
    res = []
    for file in files:
        if file.endswith(pattern):
            res.append(file)

    return res

def writeResults(file, results):
    with open(file, 'r+') as f:
        print(results, file=f)


def matchMagPhasePairs(folder):
    #create the list of all files in the folder
    files = getFiles(folder)
    phases = []
    for file in files:

        if file.endswith(('P.csv', 'PH.csv')):
            phases.append(file)
            files.remove(file)

    pairs={}
    print('phases:')
    print(phases)
    print('mags left:')
    print(files)

    for phase in phases:
        for file in files:
            name=str(file).replace('.csv', '')
            if fnmatch(phase, name+'P.csv') or fnmatch(phase, name+'PH.csv') or fnmatch(phase, name+'_PH.csv') or fnmatch(phase, name+'_P.csv'):
                pairs[phase] = file

    print('pairs:')
    print(pairs)
    return pairs


def combineAll(folder, pairs):

    headers = ['Frequency, Hz', 'Magnitude, dB', 'Phase, deg']

    for phase, mag in pairs.items():
        with open(folder + '/' + mag.replace('.csv', '') + 'COM.csv', 'w') as com:
            list=combineMagPhase(folder, phase, mag)
            com_csv = csv.writer(com)
            com_csv.writerow(headers)
            com_csv.writerows(transpose(list))

def transpose(matrix):
    return [*zip(*matrix)]

def combineMagPhase(folder, phase, mag, pairs=None):

    row_mag_ph = [[0 for x in range(0)] for y in range(3)]
    with open(folder + '/' + mag, 'rt') as mag:
        mag_csv = csv.reader(mag)
        for _ in range(18):
            headers = next(mag_csv)
        for row_mag in mag_csv:
            if row_mag[0]!='END':
                row_mag_ph[0].append(row_mag[0])
                row_mag_ph[1].append(row_mag[1])

    with open(folder + '/' + phase, 'rt') as ph:
        ph_csv = csv.reader(ph)
        for _ in range(18):
            headers = next(ph_csv)
        for row_ph in ph_csv:
            if row_ph[0] != 'END':
                row_mag_ph[2].append(row_ph[1])

    return(row_mag_ph)


def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False,
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None
    """
    from openpyxl import load_workbook

    import pandas as pd

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    # Python 2.x: define [FileNotFoundError] exception if it doesn't exist
    try:
        FileNotFoundError
    except NameError:
        FileNotFoundError = IOError

    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow,header = False, **to_excel_kwargs)

    # save the workbook
    writer.save()




def convertCSVtoTXT(csv_file, txt_file ):

    with open(txt_file, "w") as my_output_file:
        with open(csv_file, "r") as my_input_file:
            [my_output_file.write(" ".join(row) + '\n') for row in csv.reader(my_input_file)]
        my_output_file.close()
    print('conversion successful!')

# cut the data according to set frequency range
def trim(data, fmin, fmax):
    data_new = []
    fmin = fmin * 1e9  # conversion back to [Hz]
    fmax = fmax * 1e9
    for line in data:
        if len(line) > 2 and float(line[0]) > fmin and float(line[0]) < fmax:
            data_new.append(line)

    if not data_new:
        raise NotEnoughInputData('range is too narrow, list is empty!')
    return data_new


# read a csv file to buffer
def readCSV(file, headers=7):
    with open(file) as csvfile:
        reader = csv.reader(csvfile)
        data = []
        for i in range(headers):  # @UnusedVariable
            next(reader)

        for row in reader:
            data.append(row)

    print('reading of CSV file is done')
    return data


# write to txt file formatted data
def writeTXT(file, data):
    f = open(file, 'w')
    for line in data:
        if len(line) > 2:
            line = line.split(" ")
            f.write(str(line[0]) + ' ')
            f.write(str(line[1]) + ' ')
            f.write(str(line[2]) + '\n')
    f.close()
    print('writing to txt file is done')


# exception when the frequency range is chosen outside the available data
class NotEnoughInputData(Exception):
    pass


# do calibaration for dB magnitude S21
def calibrate_S21dB(data, file_cal):
    data_cal = readCSV(file_cal, headers=7)
    S21dB_cal_func = interpolate_S21dB(data_cal)
    f1, S21dB_1, phDeg_1 = format_S21(data)
    S21dB_calibrated = S21dB_1 - S21dB_cal_func(f1)
    data_calibrated = format_data(f1, S21dB_calibrated, phDeg_1)
#     plot_single(f1, S21dB_calibrated)
    return data_calibrated


# form one data array from three lists
def format_data(f, S21, phase):
    data = []
    for i in range(len(f)):
        line = str(f[i]) + ' ' + str(S21[i]) + ' ' + str(phase[i]) + '\n'
        data.append(line)
    return data


# interpolation of calibration points for its further implementation
def interpolate_S21dB(data):
    freq_Hz, S21_dB, phase = format_S21(data)  # @UnusedVariable
    S21_dB_interpol = interp1d(freq_Hz, S21_dB, kind='cubic')
    return S21_dB_interpol


# form three lists form one array
def format_S21(data):
    freq_Hz = []
    S21_dB = []
    phDeg = []
    for line in data:
        if len(line) > 2:
            freq_Hz.append(float(line[0]))
            S21_dB.append(float(line[1]))
            phDeg.append(float(line[2]))
    return freq_Hz, S21_dB, phDeg


# plot two data lists
def plot_pair(x1, y1, x2, y2):
    plt.plot(x1, y1, 'o', x2, y2, '-')
    plt.legend(['data', 'cubic'], loc='best')
    plt.show()


# plot one data list
def plot_single(x1, y1):
    plt.plot(x1, y1, 'o')
    plt.legend(['data'], loc='best')
    plt.show()
